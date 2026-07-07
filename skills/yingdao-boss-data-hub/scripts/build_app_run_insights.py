#!/usr/bin/env python3
"""Build app run insight aggregates from a tenant dashboard XLSX export."""

from __future__ import annotations

import argparse
import json
import math
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - environment guard
    print("Error: openpyxl is required. Install scripts requirements or run with the bundled Codex Python.", file=sys.stderr)
    raise


WORKSPACE_DIR = Path(__file__).resolve().parents[3]
RUNTIME_DIR = WORKSPACE_DIR / "runtime" / "yingdao-boss"
DEFAULT_OUTPUT = RUNTIME_DIR / "latest-app-run-insights.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate app run insight aggregates from exported dashboard XLSX.")
    parser.add_argument("xlsx", help="Path to exported tenant dashboard .xlsx")
    parser.add_argument("--output", "-o", default=str(DEFAULT_OUTPUT), help="Output JSON path")
    parser.add_argument("--client-name", default="", help="Client name override")
    parser.add_argument("--organization-uuid", default="", help="Organization UUID override")
    parser.add_argument("--custom-no", default="", help="Client custom number override")
    parser.add_argument("--tenant-uuid", default="", help="Tenant UUID override")
    return parser.parse_args()


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def to_seconds(value: Any) -> int:
    if value is None:
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def parse_dt(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    text = clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).replace(microsecond=0)
    except ValueError:
        return None


def fmt_dt(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def fmt_date(value: datetime | None) -> str:
    return value.strftime("%Y%m%d") if value else ""


def row_dicts(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [clean(v) for v in next(rows)]
    except StopIteration:
        return []
    result = []
    for row in rows:
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        if any(v is not None and clean(v) for v in item.values()):
            result.append(item)
    return result


def update_aggregate(target: dict[str, Any], *, seconds: int, start: datetime, app_uuid: str, runner_uuid: str) -> None:
    target["runCnt"] += 1
    target["durationSeconds"] += seconds
    target["activeDates"].add(fmt_date(start))
    target["appUuids"].add(app_uuid)
    target["runnerUuids"].add(runner_uuid)
    if not target["firstRunAt"] or start < target["firstRunAt"]:
        target["firstRunAt"] = start
    if not target["lastRunAt"] or start > target["lastRunAt"]:
        target["lastRunAt"] = start


def base_aggregate() -> dict[str, Any]:
    return {
        "runCnt": 0,
        "durationSeconds": 0,
        "activeDates": set(),
        "appUuids": set(),
        "runnerUuids": set(),
        "firstRunAt": None,
        "lastRunAt": None,
    }


def finalize_aggregate(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "runCnt": item["runCnt"],
        "durationSeconds": item["durationSeconds"],
        "activeDays": len(item["activeDates"]),
        "appCount": len(item["appUuids"]),
        "runnerCount": len(item["runnerUuids"]),
        "firstRunAt": fmt_dt(item["firstRunAt"]),
        "lastRunAt": fmt_dt(item["lastRunAt"]),
    }


def app_display_name(app_uuid: str, app_name: str) -> str:
    if app_name:
        return app_name
    suffix = app_uuid[-8:] if app_uuid else "unknown"
    return f"未命名应用 {suffix}"


def build_document(args: argparse.Namespace) -> dict[str, Any]:
    source = Path(args.xlsx).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(source)

    workbook = load_workbook(source, read_only=True, data_only=True)
    detail_rows = row_dicts(workbook, "运行明细数据")
    user_rows = row_dicts(workbook, "用户汇总数据")
    enterprise_rows = row_dicts(workbook, "企业汇总数据")

    user_map: dict[str, dict[str, str]] = {}
    for row in user_rows:
        uuid = clean(row.get("用户uuid"))
        if not uuid:
            continue
        user_map[uuid] = {
            "runnerName": clean(row.get("用户名称")),
            "runnerAccount": clean(row.get("用户账号")),
            "runnerStatus": clean(row.get("用户状态")),
            "runnerAccountType": clean(row.get("账号类型")),
        }

    client_name = args.client_name
    range_start = ""
    range_end = ""
    if enterprise_rows:
        enterprise = enterprise_rows[0]
        client_name = client_name or clean(enterprise.get("企业名称"))
        range_start = clean(enterprise.get("数据范围开始时间"))
        range_end = clean(enterprise.get("数据范围结束时间"))

    apps: dict[str, dict[str, Any]] = defaultdict(base_aggregate)
    runners: dict[str, dict[str, Any]] = defaultdict(base_aggregate)
    app_runner: dict[tuple[str, str], dict[str, Any]] = defaultdict(base_aggregate)
    app_daily: dict[tuple[str, str], dict[str, Any]] = defaultdict(base_aggregate)
    runner_daily: dict[tuple[str, str], dict[str, Any]] = defaultdict(base_aggregate)
    app_runner_daily: dict[tuple[str, str, str], dict[str, Any]] = defaultdict(base_aggregate)
    app_names: dict[str, str] = {}
    runner_accounts: dict[str, str] = {}
    runner_names: dict[str, str] = {}
    runner_types: dict[str, str] = {}
    min_start: datetime | None = None
    max_start: datetime | None = None

    for row in detail_rows:
        app_uuid = clean(row.get("应用uuid"))
        runner_uuid = clean(row.get("运行者uuid"))
        start = parse_dt(row.get("运行开始时间"))
        if not app_uuid or not runner_uuid or not start:
            continue

        app_name = clean(row.get("应用名称"))
        runner_account = clean(row.get("运行者账号")) or user_map.get(runner_uuid, {}).get("runnerAccount", "")
        runner_name = user_map.get(runner_uuid, {}).get("runnerName", "") or runner_account
        runner_type = clean(row.get("运行者账号类型")) or user_map.get(runner_uuid, {}).get("runnerAccountType", "")
        seconds = to_seconds(row.get("运行时长（秒）"))
        date = fmt_date(start)

        app_names.setdefault(app_uuid, app_name)
        if app_name and not app_names[app_uuid]:
            app_names[app_uuid] = app_name
        runner_accounts.setdefault(runner_uuid, runner_account)
        runner_names.setdefault(runner_uuid, runner_name)
        runner_types.setdefault(runner_uuid, runner_type)

        for target in (
            apps[app_uuid],
            runners[runner_uuid],
            app_runner[(app_uuid, runner_uuid)],
            app_daily[(app_uuid, date)],
            runner_daily[(runner_uuid, date)],
            app_runner_daily[(app_uuid, runner_uuid, date)],
        ):
            update_aggregate(target, seconds=seconds, start=start, app_uuid=app_uuid, runner_uuid=runner_uuid)

        min_start = start if min_start is None else min(min_start, start)
        max_start = start if max_start is None else max(max_start, start)

    if not range_start and min_start:
        range_start = fmt_date(min_start)
    if not range_end and max_start:
        range_end = fmt_date(max_start)

    app_items = []
    for app_uuid, agg in apps.items():
        item = finalize_aggregate(agg)
        item.update({
            "appUuid": app_uuid,
            "appName": app_display_name(app_uuid, app_names.get(app_uuid, "")),
        })
        app_items.append(item)

    runner_items = []
    for runner_uuid, agg in runners.items():
        item = finalize_aggregate(agg)
        account = runner_accounts.get(runner_uuid, "")
        item.update({
            "runnerUuid": runner_uuid,
            "runnerName": runner_names.get(runner_uuid) or account,
            "runnerAccount": account,
            "runnerAccountType": runner_types.get(runner_uuid, ""),
        })
        runner_items.append(item)

    pair_items = []
    for (app_uuid, runner_uuid), agg in app_runner.items():
        item = finalize_aggregate(agg)
        item.update({
            "appUuid": app_uuid,
            "appName": app_display_name(app_uuid, app_names.get(app_uuid, "")),
            "runnerUuid": runner_uuid,
            "runnerName": runner_names.get(runner_uuid) or runner_accounts.get(runner_uuid, ""),
            "runnerAccount": runner_accounts.get(runner_uuid, ""),
        })
        pair_items.append(item)

    app_daily_items = []
    for (app_uuid, date), agg in app_daily.items():
        item = finalize_aggregate(agg)
        item.update({
            "date": date,
            "appUuid": app_uuid,
            "appName": app_display_name(app_uuid, app_names.get(app_uuid, "")),
        })
        app_daily_items.append(item)

    runner_daily_items = []
    for (runner_uuid, date), agg in runner_daily.items():
        item = finalize_aggregate(agg)
        item.update({
            "date": date,
            "runnerUuid": runner_uuid,
            "runnerName": runner_names.get(runner_uuid) or runner_accounts.get(runner_uuid, ""),
            "runnerAccount": runner_accounts.get(runner_uuid, ""),
        })
        runner_daily_items.append(item)

    app_runner_daily_items = []
    for (app_uuid, runner_uuid, date), agg in app_runner_daily.items():
        item = finalize_aggregate(agg)
        item.update({
            "date": date,
            "appUuid": app_uuid,
            "appName": app_display_name(app_uuid, app_names.get(app_uuid, "")),
            "runnerUuid": runner_uuid,
            "runnerName": runner_names.get(runner_uuid) or runner_accounts.get(runner_uuid, ""),
            "runnerAccount": runner_accounts.get(runner_uuid, ""),
        })
        app_runner_daily_items.append(item)

    sort_key = lambda item: (-item["runCnt"], -item["durationSeconds"], item.get("appName") or item.get("runnerAccount") or "")
    app_items.sort(key=sort_key)
    runner_items.sort(key=sort_key)
    pair_items.sort(key=sort_key)
    app_daily_items.sort(key=lambda item: (item["date"], -item["runCnt"], item["appName"]))
    runner_daily_items.sort(key=lambda item: (item["date"], -item["runCnt"], item["runnerAccount"]))
    app_runner_daily_items.sort(key=lambda item: (item["date"], item["appUuid"], -item["runCnt"]))

    total_seconds = sum(item["durationSeconds"] for item in app_items)
    total_runs = sum(item["runCnt"] for item in app_items)
    active_dates = sorted({date for _, date in app_daily.keys()})

    return {
        "schema": "yingdao-boss-app-run-insights.v1",
        "meta": {
            "fetched_at": datetime.now().astimezone().isoformat(),
            "client_name": client_name,
            "organization_uuid": args.organization_uuid,
            "custom_no": args.custom_no,
            "tenant_uuid": args.tenant_uuid,
            "source_file": str(source),
            "range_start": range_start,
            "range_end": range_end,
            "run_detail_available": bool(detail_rows),
            "run_detail_rows": len(detail_rows),
        },
        "client_summary": {
            "totalRunCnt": total_runs,
            "totalRunDurationSeconds": total_seconds,
            "appCount": len(app_items),
            "runnerCount": len(runner_items),
            "appRunnerPairCount": len(pair_items),
            "activeDays": len(active_dates),
            "firstRunAt": fmt_dt(min_start),
            "lastRunAt": fmt_dt(max_start),
        },
        "apps": app_items,
        "runners": runner_items,
        "app_runner_matrix": pair_items,
        "runner_daily_activity": runner_daily_items,
        "app_daily_summary": app_daily_items,
        "app_runner_daily_activity": app_runner_daily_items,
    }


def main() -> int:
    args = parse_args()
    try:
        doc = build_document(args)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote app run insights to {output}")
    print(
        "  "
        + f"{doc['client_summary']['totalRunCnt']} runs, "
        + f"{doc['client_summary']['appCount']} apps, "
        + f"{doc['client_summary']['runnerCount']} runners"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
